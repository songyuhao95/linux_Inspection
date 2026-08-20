"""tests/test_render_xlsx.py — T-106 Excel 渲染层测试（合同 AC-1..AC-2）。

覆盖（合同必需步骤 3 + mitigations）：
  - 三 Sheet（Overview/Local/Errors-Evidence）与 RR §3 列/内容（模块级
    常量 + 夹具样本结构断言，xlsxwriter 缺失环境下仍可运行）；
  - 四状态色板常量（RR §5：#2E7D32/#F9A825/#C62828/#757575）；
  - 状态计数与 JSON execution_summary 一致；UNKNOWN 独立计数、不混入 OK；
  - 文件名 `<inspection-id>.xlsx` 与 out_path（CLI `--excel PATH` 语义）覆盖；
  - xlsxwriter 缺失 → RendererError(exit_code=10)（TD §8 负向测试，
    通过 sys.modules 置 None 在任意环境确定性触发，缺省环境可运行）；
  - 损坏/结构不完整 JSON → RendererError(exit_code=10)；
  - 实际渲染断言（xlsxwriter 可用时，openpyxl 只读校验，TD §8）：
    三 Sheet 存在、状态计数与 JSON 一致、CRIT 值红色字体、status 背景色、
    Errors-Evidence 行集合；xlsxwriter 缺失环境 skip 渲染类，保留常量/
    结构断言（合同约束：本任务不安装依赖）。

只读使用 tests/fixtures/json/（T-104 夹具）与 tests/fixtures/xlsx/
（本任务夹具样本）；不连接、不执行命令、不安装依赖。
"""

import json
import sys
from pathlib import Path

import pytest

from inspect import render_xlsx as rx

FIXTURE_JSON = Path(__file__).parent / "fixtures" / "json"
FIXTURE_XLSX = Path(__file__).parent / "fixtures" / "xlsx"
VALID_FIXTURE = FIXTURE_JSON / "host-result-valid.json"

_RUN_ID = "run-20260814-001"
_INSP_ID = "insp-20260814120000-node-fx01"


def ok_doc() -> dict:
    """T-104 夹具：单主机单指标（swap OK，SUCCESS）。"""
    return json.loads(VALID_FIXTURE.read_text(encoding="utf-8"))


def make_metric(
    metric_id: str,
    status: str,
    raw_value,
    normalized_value,
    unit: str = "%",
    *,
    name: str = None,
    layer: str = "document-baseline",
    rule_id: str = None,
    value=None,
    source_anchor: str = "测试来源锚点",
    threshold_notes: str = None,
    ev_command: str = "echo test",
    ev_summary: str = "out",
    error: dict = None,
    prov_notes: str = None,
    ev_details: list = None,
) -> dict:
    """构造符合 host-result-v1 schema 的 metric 对象（HR §3.1）。"""
    metric = {
        "metric_id": metric_id,
        "name": name or metric_id,
        "scope": "local-common-p0-v1",
        "status": status,
        "raw_value": raw_value,
        "normalized_value": normalized_value,
        "unit": unit,
        "threshold": {
            "layer": layer,
            "rule_id": rule_id,
            "value": value,
            "source_anchor": source_anchor,
            "notes": threshold_notes,
        },
        "evidence": {
            "command": ev_command,
            "output_summary": ev_summary,
            "sampled_at": "2026-08-14T12:00:00+08:00",
        },
        "error": error,
        "provenance": {
            "config_sources": [],
            "doc_sources": ["测试巡检手册 v1"],
            "notes": prov_notes,
        },
    }
    if ev_details is not None:
        metric["evidence"]["details"] = ev_details
    return metric


def make_doc(
    host_name: str,
    metrics: list,
    *,
    inspection_id: str = _INSP_ID,
    execution_status: str = "SUCCESS",
) -> dict:
    """按 metrics 构造完整 host-result-v1 文档（execution_summary 自算，
    与 T-104 normalize 口径一致：error 非空 → status=UNKNOWN）。"""
    return {
        "schema": "host-result-v1",
        "schema_version": 1,
        "run_id": _RUN_ID,
        "inspection_id": inspection_id,
        "host": {
            "name": host_name,
            "ip": "<IP>",
            "inventory_source": "local",
            "product_profiles": ["elasticsearch"],
        },
        "collected_at": "2026-08-14T12:00:00+08:00",
        "duration_sec": 12.34,
        "execution_status": execution_status,
        "execution_summary": {
            "total_metrics": len(metrics),
            "ok": sum(1 for m in metrics if m["status"] == "OK"),
            "warn": sum(1 for m in metrics if m["status"] == "WARN"),
            "crit": sum(1 for m in metrics if m["status"] == "CRIT"),
            "unknown": sum(1 for m in metrics if m["status"] == "UNKNOWN"),
            "executed": len(metrics),
            "failed": sum(1 for m in metrics if m.get("error") is not None),
        },
        "metrics": metrics,
        "meta": {
            "control_endpoint": "Linux/WSL Python3",
            "gather_facts": False,
            "serial": 1,
            "become_scope": "minimal",
            "generator": "inspect.sh",
            "generator_version": "0.1.0-draft",
        },
    }


def rich_docs() -> list:
    """两台主机，覆盖 OK/WARN/CRIT/UNKNOWN + 技术错误（error 非空）+
    文档冲突 UNKNOWN（error 为空），合计 7 指标。

    node-fx01：OK=1 WARN=1 CRIT=1 UNKNOWN=2（含 1 个 error=PERMISSION_DENIED）
    node-fx02：OK=1 UNKNOWN=1（C5 等级缺失，error 为空）
    合计：OK=2 WARN=1 CRIT=1 UNKNOWN=3
    """
    node1 = make_doc("node-fx01", [
        make_metric(
            "local.cpu.utilization", "OK", "35.2", 35.2,
            rule_id="linux-common-p0-v1.cpu.utilization.ok", value="<70",
        ),
        make_metric(
            "local.memory.available_percent", "WARN", "12", 12.0,
            rule_id="linux-common-p0-v1.memory.available_percent.warn",
            value="10-20 缺失", prov_notes="C4 缺失边界",
        ),
        make_metric(
            "local.filesystem.used_percent", "CRIT", "91", 91.0,
            rule_id="linux-common-p0-v1.filesystem.used_percent.crit",
            value=">85", source_anchor="Mysql 手册 T5R9", ev_summary="91%",
        ),
        make_metric(
            "local.swap.used_percent", "UNKNOWN", "12.5", 12.5,
            layer="unresolved-document-conflict", rule_id=None, value=None,
            source_anchor="9 份手册 P0 内存行（C3）",
            prov_notes="文档冲突 C3 unresolved",
        ),
        make_metric(
            "local.logs.key_evidence", "UNKNOWN", None, None, unit="count",
            layer=None, rule_id=None, value=None, source_anchor=None,
            ev_command="tail -300 /opt/redis/logs/redis.log",
            ev_summary=None,
            error={
                "code": "PERMISSION_DENIED",
                "message": "cannot read /opt/redis/logs/redis.log (permission denied)",
                "metric_status": "UNKNOWN",
            },
            prov_notes="权限不足，指标 UNKNOWN",
        ),
    ])
    node2 = make_doc("node-fx02", [
        make_metric(
            "local.filesystem.inode_used_percent", "OK", "41", 41.0,
            rule_id="linux-common-p0-v1.filesystem.inode_used_percent.ok",
            value="<80",
        ),
        make_metric(
            "local.cpu.load_1m", "UNKNOWN", "2.5", 2.5,
            layer="unresolved-document-conflict", rule_id=None, value=None,
            source_anchor="C5 等级缺失", prov_notes="等级缺失 C5",
        ),
    ])
    return [node1, node2]


def load_workbook(path: Path):
    """openpyxl 只读校验（TD §8：dev 依赖仅校验）。"""
    from openpyxl import load_workbook as _load

    return _load(path)


def find_label_row(ws, label: str) -> int:
    """Overview 中 A 列标签所在行号（1-based）。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return cell.row
    raise AssertionError(f"Overview 缺少标签: {label}")


def local_rows(ws):
    """Local 数据行（跳过表头），每行 dict（列名 → 值）。"""
    out = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        out.append(dict(zip(rx.LOCAL_HEADERS, values)))
    return out


def errors_rows(ws):
    """Errors-Evidence 数据行（跳过表头），每行 dict。"""
    out = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        out.append(dict(zip(rx.ERRORS_HEADERS, values)))
    return out


# --------------------------------------------------------------------------
# 1. 模块级常量/结构契约（RR §3/§5）——xlsxwriter 缺失环境下仍可运行
# --------------------------------------------------------------------------


class TestConstants:
    def test_status_colors_match_rr_section5(self):
        assert rx.COLOR_OK == "#2E7D32"
        assert rx.COLOR_WARN == "#F9A825"
        assert rx.COLOR_CRIT == "#C62828"
        assert rx.COLOR_UNKNOWN == "#757575"
        assert rx.STATUS_COLORS == {
            "OK": "#2E7D32", "WARN": "#F9A825",
            "CRIT": "#C62828", "UNKNOWN": "#757575",
        }

    def test_sheet_names_match_rr_section3(self):
        assert rx.SHEET_NAMES == ("Overview", "Local", "nginx", "keepalived", "Errors-Evidence")
        assert rx.SHEET_NGINX == "nginx"
        assert rx.NGINX_HEADERS == rx.LOCAL_HEADERS

    def test_local_headers_cover_rr_section3_columns(self):
        assert rx.LOCAL_HEADERS[:2] == ("host", "ip")
        for col in (
            "metric_id", "raw_value", "normalized_value", "unit", "status",
            "threshold_rule", "command",
        ):
            assert col in rx.LOCAL_HEADERS
        for removed in ("source_anchor", "evidence_summary", "provenance"):
            assert removed not in rx.LOCAL_HEADERS

    def test_errors_headers_cover_rr_section3_columns(self):
        for col in ("error_code", "message", "command", "output_summary"):
            assert col in rx.ERRORS_HEADERS

    def test_renderer_error_exit_code_10(self):
        assert rx.EXIT_RENDER_ERROR == 10
        err = rx.RendererError("x")
        assert err.exit_code == 10


# --------------------------------------------------------------------------
# 2. 状态计数：与 JSON 一致、UNKNOWN 不混入 OK（RR §6）
# --------------------------------------------------------------------------


class TestCountStatuses:
    def test_ok_fixture_counts(self):
        doc = ok_doc()
        assert rx.count_statuses(doc) == {"OK": 1, "WARN": 0, "CRIT": 0, "UNKNOWN": 0}

    def test_rich_docs_unknown_not_mixed_into_ok(self):
        doc = rich_docs()[0]
        counts = rx.count_statuses(doc)
        assert counts == {"OK": 1, "WARN": 1, "CRIT": 1, "UNKNOWN": 2}
        # UNKNOWN 独立计数：四状态合计 == metrics 总数，OK 不含 UNKNOWN
        assert counts["OK"] == 1
        assert sum(counts.values()) == len(doc["metrics"])

    def test_counts_match_execution_summary(self):
        for doc in rich_docs():
            counts = rx.count_statuses(doc)
            summary = doc["execution_summary"]
            assert counts["OK"] == summary["ok"]
            assert counts["WARN"] == summary["warn"]
            assert counts["CRIT"] == summary["crit"]
            assert counts["UNKNOWN"] == summary["unknown"]

    def test_invalid_status_raises(self):
        doc = ok_doc()
        doc["metrics"][0]["status"] = "FATAL"
        with pytest.raises(rx.RendererError, match="FATAL"):
            rx.count_statuses(doc)


# --------------------------------------------------------------------------
# 3. 结构校验与损坏检测（RendererError，exit_code=10 语义）
# --------------------------------------------------------------------------


class TestValidation:
    def test_validate_ok_fixture(self):
        rx.validate_doc(ok_doc())

    def test_missing_top_level_key_raises(self):
        doc = ok_doc()
        del doc["execution_status"]
        with pytest.raises(rx.RendererError) as exc:
            rx.validate_doc(doc)
        assert exc.value.exit_code == 10
        assert "execution_status" in str(exc.value)

    def test_missing_metric_key_raises(self):
        doc = ok_doc()
        del doc["metrics"][0]["provenance"]
        with pytest.raises(rx.RendererError, match="provenance"):
            rx.validate_doc(doc)

    def test_invalid_metric_status_raises(self):
        doc = ok_doc()
        doc["metrics"][0]["status"] = "FATAL"
        with pytest.raises(rx.RendererError, match="FATAL"):
            rx.validate_doc(doc)

    def test_error_without_unknown_status_raises(self):
        # HR §3.2：error 存在 → status 一律 UNKNOWN，不得伪装业务状态
        doc = ok_doc()
        doc["metrics"][0]["status"] = "CRIT"
        doc["metrics"][0]["error"] = {
            "code": "TIMEOUT", "message": "timeout", "metric_status": "UNKNOWN",
        }
        with pytest.raises(rx.RendererError, match="UNKNOWN"):
            rx.validate_doc(doc)

    def test_count_mismatch_raises(self):
        # RR §6：状态计数与 JSON 不一致 → 明确报错，不静默呈现
        doc = ok_doc()
        doc["execution_summary"]["ok"] += 1
        with pytest.raises(rx.RendererError, match="不一致"):
            rx.validate_doc(doc)

    def test_empty_docs_raises(self):
        with pytest.raises(rx.RendererError, match="没有可渲染"):
            rx.render_xlsx([])

    def test_cross_doc_inspection_id_mismatch_raises(self):
        doc1, doc2 = rich_docs()
        doc2["inspection_id"] = "insp-other-999"
        with pytest.raises(rx.RendererError, match="inspection_id"):
            rx.render_xlsx([doc1, doc2])


# --------------------------------------------------------------------------
# 4. 负向测试：xlsxwriter 缺失 → 明确报错退出码 10（TD §8，任意环境可运行）
# --------------------------------------------------------------------------


class TestMissingXlsxwriter:
    def test_missing_xlsxwriter_raises_renderer_error(self, monkeypatch, tmp_path):
        # 确定性模拟缺失：sys.modules["xlsxwriter"]=None → import 抛 ImportError
        monkeypatch.setitem(sys.modules, "xlsxwriter", None)
        with pytest.raises(rx.RendererError) as exc:
            rx.render_xlsx(ok_doc(), tmp_path / "x.xlsx")
        assert exc.value.exit_code == 10
        assert "xlsxwriter" in str(exc.value)
        assert not (tmp_path / "x.xlsx").exists()

    def test_missing_xlsxwriter_via_file_entry(self, monkeypatch, tmp_path):
        monkeypatch.setitem(sys.modules, "xlsxwriter", None)
        with pytest.raises(rx.RendererError) as exc:
            rx.render_xlsx_file(VALID_FIXTURE, tmp_path / "x.xlsx")
        assert exc.value.exit_code == 10


# --------------------------------------------------------------------------
# 5. 文件入口与损坏检测（不依赖 xlsxwriter 的错误路径）
# --------------------------------------------------------------------------


class TestRenderXlsxFileErrors:
    def test_corrupt_json_raises(self, tmp_path):
        bad = tmp_path / "bad.json"
        bad.write_text("{not json", encoding="utf-8")
        with pytest.raises(rx.RendererError) as exc:
            rx.render_xlsx_file(bad, tmp_path / "x.xlsx")
        assert exc.value.exit_code == 10
        assert "JSON" in str(exc.value)

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(rx.RendererError) as exc:
            rx.render_xlsx_file(tmp_path / "nope.json")
        assert exc.value.exit_code == 10

    def test_wrong_json_shape_raises(self, tmp_path):
        bad = tmp_path / "bad.json"
        bad.write_text(json.dumps(42), encoding="utf-8")
        with pytest.raises(rx.RendererError, match="结构非法"):
            rx.render_xlsx_file(bad)


# --------------------------------------------------------------------------
# 6. 夹具样本结构（openpyxl 只读校验；xlsxwriter 缺失环境的结构断言）
# --------------------------------------------------------------------------


class TestFixtureSample:
    def test_sample_sheets_and_headers(self):
        wb = load_workbook(FIXTURE_XLSX / "host-result-valid.xlsx")
        # 该检查入库样例为 nginx Sheet 引入前的旧版三 Sheet 工作簿；
        # 新 Sheet 布局在 test_sheets_registered_and_counts 用 stub 校验。
        assert wb.sheetnames == ["Overview", "Local", "Errors-Evidence"]
        assert list(rx.SHEET_NAMES) == ["Overview", "Local", "nginx", "keepalived", "Errors-Evidence"]
        # This checked-in workbook is a legacy sample; new headers are tested
        # against freshly rendered output below without rewriting the fixture.
        legacy_headers = [c.value for c in wb[rx.SHEET_LOCAL][1]]
        assert legacy_headers[0] == "host"
        assert "source_anchor" in legacy_headers
        assert [c.value for c in wb[rx.SHEET_ERRORS][1]] == list(rx.ERRORS_HEADERS)

    def test_sample_overview_contents(self):
        wb = load_workbook(FIXTURE_XLSX / "host-result-valid.xlsx")
        ws = wb[rx.SHEET_OVERVIEW]
        # 仅取 run 信息区（"主机×状态汇总"之前的行；该表头行同为
        # host 标签，不得混入 run 信息）
        section = find_label_row(ws, "主机×状态汇总")
        values = {}
        for row in ws.iter_rows():
            if row[0].row >= section:
                break
            label, value = row[0].value, row[1].value
            if label and value is not None:
                values[label] = value
        assert values["inspection_id"] == _INSP_ID
        assert values["run_id"] == _RUN_ID
        assert values["execution_status"] == "SUCCESS"
        assert values["host"] == "node-fx01"
        assert values["阈值版本"] == "linux-common-p0-v1"
        assert values["生成时间"]

    def test_sample_local_status_fill_uses_rr_color(self):
        # 夹具样本 Local 中 OK 行 status 单元格为 RR §5 绿色 #2E7D32
        wb = load_workbook(FIXTURE_XLSX / "host-result-valid.xlsx")
        ws = wb[rx.SHEET_LOCAL]
        status_cell = ws.cell(2, 7)  # status 列
        assert status_cell.value == "OK"
        rgb = status_cell.fill.start_color.rgb
        assert rgb and rgb[-6:] == "2E7D32"


# --------------------------------------------------------------------------
# 7. 桩 xlsxwriter 渲染逻辑验证（缺省环境可运行：不依赖真实库，
#    验证三 Sheet 注册/状态样式/CRIT 值红色字体/UNKNOWN 不混入 OK）
# --------------------------------------------------------------------------


class StubFormat:
    def __init__(self, **kwargs):
        self.kwargs = kwargs


class StubWorksheet:
    def __init__(self, name):
        self.name = name
        self.cells = {}  # (row, col) -> (value, format)
        self.widths = {}
        self.frozen = None

    def write(self, row, col, value, fmt=None):
        self.cells[(row, col)] = (value, fmt)

    def set_column(self, first, last, width):
        self.widths[(first, last)] = width

    def freeze_panes(self, row, col):
        self.frozen = (row, col)


class StubWorkbook:
    def __init__(self, path):
        self.path = str(path)
        self.sheets = []
        self.formats = []
        self.closed = False

    def add_format(self, properties=None, **kwargs):
        # xlsxwriter 语义：位置 dict 与关键字均可
        fmt = StubFormat(**(properties or {}), **kwargs)
        self.formats.append(fmt)
        return fmt

    def add_worksheet(self, name):
        ws = StubWorksheet(name)
        self.sheets.append(ws)
        return ws

    def close(self):
        self.closed = True


class StubXlsxwriter:
    def __init__(self):
        self.workbooks = []

    def Workbook(self, path):
        wb = StubWorkbook(path)
        self.workbooks.append(wb)
        return wb


def stub_label_row(ws, label):
    for (row, col), (value, _fmt) in ws.cells.items():
        if col == 0 and value == label:
            return row
    raise AssertionError(f"桩工作表缺少标签: {label}")


class TestRenderWithStubXlsxwriter:
    def test_sheets_registered_and_counts(self, monkeypatch, tmp_path):
        stub = StubXlsxwriter()
        monkeypatch.setitem(sys.modules, "xlsxwriter", stub)
        out = rx.render_xlsx(rich_docs(), tmp_path / "r.xlsx")
        assert out == tmp_path / "r.xlsx"
        wb = stub.workbooks[-1]
        assert wb.closed
        assert [s.name for s in wb.sheets] == list(rx.SHEET_NAMES)
        overview = wb.sheets[0]

        # run 信息（RR §3）
        def value_of(label):
            return overview.cells[(stub_label_row(overview, label), 1)][0]

        assert value_of("inspection_id") == _INSP_ID
        assert value_of("host") == "node-fx01, node-fx02"
        assert value_of("execution_status") == "SUCCESS"
        assert value_of("阈值版本") == "linux-common-p0-v1"
        assert value_of("生成时间")

        # 状态计数（合计）：UNKNOWN 独立计数、不混入 OK（RR §3/§5）
        section = stub_label_row(overview, "状态计数（合计）")
        counts_row = section + 2
        ok_v = overview.cells[(counts_row, 0)][0]
        warn_v = overview.cells[(counts_row, 1)][0]
        crit_v = overview.cells[(counts_row, 2)][0]
        unk_v = overview.cells[(counts_row, 3)][0]
        total = overview.cells[(counts_row, 4)][0]
        assert (ok_v, warn_v, crit_v, unk_v, total) == (2, 1, 1, 3, 7)

    def test_status_styles_and_crit_red_font(self, monkeypatch, tmp_path):
        stub = StubXlsxwriter()
        monkeypatch.setitem(sys.modules, "xlsxwriter", stub)
        rx.render_xlsx(rich_docs(), tmp_path / "r.xlsx")
        wb = stub.workbooks[-1]
        local = wb.sheets[1]

        # Local：每主机每指标一行（7 行数据），status 单元格带 RR §5 背景色
        rows = {
            row: {
                col: (value, fmt)
                for (r, col), (value, fmt) in local.cells.items()
                if r == row
            }
            for (row, _col) in local.cells
            if row >= 1
        }
        statuses = {
            row: cols[7][0]
            for row, cols in rows.items()
            if 7 in cols
        }
        assert sorted(statuses.values()) == ["CRIT", "OK", "OK", "UNKNOWN",
                                             "UNKNOWN", "UNKNOWN", "WARN"]
        for row, cols in rows.items():
            status = cols[7][0]
            fmt = cols[7][1]
            assert fmt.kwargs["bg_color"] == rx.STATUS_COLORS[status]

        # CRIT 值红色字体（用户需求）：raw_value/normalized_value 列
        crit_row = next(row for row, cols in rows.items() if cols[7][0] == "CRIT")
        for col in (4, 5):
            fmt = rows[crit_row][col][1]
            assert fmt.kwargs.get("font_color") == rx.COLOR_CRIT
        # 非 CRIT 行值单元格不套红
        ok_row = next(row for row, cols in rows.items() if cols[7][0] == "OK")
        assert rows[ok_row][4][1].kwargs.get("font_color") != rx.COLOR_CRIT

        # 新 Local 列：ip、清晰阈值解释、事实源 command；旧冗余列删除。
        assert rows[crit_row][1][0] == "<IP>"
        assert rows[crit_row][8][0]
        assert "判定规则：" in rows[crit_row][8][0]
        assert rows[crit_row][9][0] == "echo test"

    def test_host_ips_override_fact_source_ip_for_local_only(self, monkeypatch, tmp_path):
        stub = StubXlsxwriter()
        monkeypatch.setitem(sys.modules, "xlsxwriter", stub)
        docs = rich_docs()
        host_ips = {
            "node-fx01": "192.0.2.101",
            "node-fx02": "192.0.2.102",
        }

        rx.render_xlsx(docs, tmp_path / "r.xlsx", host_ips=host_ips)

        local = stub.workbooks[-1].sheets[1]
        rows = [
            local.cells[(row, 1)][0]
            for row in range(1, 1 + sum(len(rx._metric_rows(m)) for d in docs for m in d["metrics"]))
        ]
        assert rows == [
            "192.0.2.101", "192.0.2.101", "192.0.2.101", "192.0.2.101",
            "192.0.2.101", "192.0.2.102", "192.0.2.102",
        ]
        assert all(doc["host"]["ip"] == "<IP>" for doc in docs)


    def test_local_expands_load_and_filesystem_details(self, monkeypatch, tmp_path):
        stub = StubXlsxwriter()
        monkeypatch.setitem(sys.modules, "xlsxwriter", stub)
        details_doc = make_doc("node-detail", [
            make_metric(
                "local.cpu.load_1m", "OK", "0.5", 0.5,
                name="系统负载", value="load <= CPU cores",
                ev_command="cat /proc/loadavg; nproc",
                ev_details=[
                    {"window": "1 分钟", "load": 0.5, "cpu_cores": 2,
                     "status": "OK", "judgement": "正常"},
                    {"window": "5 分钟", "load": 0.4, "cpu_cores": 2,
                     "status": "OK", "judgement": "正常"},
                    {"window": "15 分钟", "load": 0.3, "cpu_cores": 2,
                     "status": "OK", "judgement": "正常"},
                ],
            ),
            make_metric(
                "local.filesystem.used_percent", "CRIT", "100", 100.0,
                name="磁盘使用率", value=">85",
                ev_command="df -hT",
                ev_details=[
                    {"filesystem": "/dev/root", "mount": "/", "used_percent": 61.0,
                     "status": "OK"},
                    {"filesystem": "/dev/iso", "mount": "/mnt/iso", "used_percent": 100.0,
                     "status": "CRIT"},
                ],
            ),
            make_metric(
                "local.filesystem.inode_used_percent", "OK", "1", 1.0,
                name="inode 使用率", value="<80",
                ev_command="df -i",
                ev_details=[
                    {"filesystem": "/dev/root", "mount": "/", "used_percent": 1.0,
                     "status": "OK"},
                    {"filesystem": "/dev/iso", "mount": "/mnt/iso", "used_percent": 1.0,
                     "status": "OK"},
                ],
            ),
        ])
        rx.render_xlsx(details_doc, tmp_path / "r.xlsx")
        local = stub.workbooks[-1].sheets[1]
        rows = []
        for row in range(1, 8):
            rows.append({
                rx.LOCAL_HEADERS[col]: local.cells[(row, col)][0]
                for col in range(len(rx.LOCAL_HEADERS))
            })
        assert len(rows) == 7
        assert [row["ip"] for row in rows] == ["<IP>"] * 7
        assert [row["name"] for row in rows[:3]] == [
            "1 分钟系统负载", "5 分钟系统负载", "15 分钟系统负载",
        ]
        assert [row["name"] for row in rows[3:5]] == [
            "磁盘使用率: /", "磁盘使用率: /mnt/iso",
        ]
        assert [row["name"] for row in rows[5:]] == [
            "inode 使用率: /", "inode 使用率: /mnt/iso",
        ]
        assert [row["raw_value"] for row in rows[:3]] == [0.5, 0.4, 0.3]
        assert [row["raw_value"] for row in rows[3:5]] == [61.0, 100.0]
        assert [row["status"] for row in rows] == [
            "OK", "OK", "OK", "OK", "CRIT", "OK", "OK",
        ]
        assert all(row["command"] == "cat /proc/loadavg; nproc" for row in rows[:3])
        assert all(row["command"] == "df -hT" for row in rows[3:5])
        assert all(row["command"] == "df -i" for row in rows[5:])
        assert "测量周期：1 分钟" in rows[0]["threshold_rule"]
        assert "判定规则：load ≤ CPU cores" in rows[0]["threshold_rule"]
        assert "挂载点：/mnt/iso" in rows[4]["threshold_rule"]

    def test_errors_evidence_rows(self, monkeypatch, tmp_path):
        stub = StubXlsxwriter()
        monkeypatch.setitem(sys.modules, "xlsxwriter", stub)
        rx.render_xlsx(rich_docs(), tmp_path / "r.xlsx")
        wb = stub.workbooks[-1]
        errors = wb.sheets[4]  # Overview(0) Local(1) nginx(2) keepalived(3) Errors-Evidence(4)

        # error 非空指标（PERMISSION_DENIED）+ 文档冲突/缺失 UNKNOWN 清单
        metric_ids = {
            errors.cells[(row, 1)][0]
            for (row, col) in errors.cells
            if col == 1 and row >= 1
        }
        assert metric_ids == {
            "local.logs.key_evidence", "local.swap.used_percent",
            "local.cpu.load_1m",
        }
        # error 行：code/message/command/output_summary（RR §3）
        err_rows = [
            row for (row, col) in errors.cells
            if col == 4 and row >= 1 and errors.cells[(row, col)][0] == "PERMISSION_DENIED"
        ]
        assert len(err_rows) == 1
        err_row = err_rows[0]
        assert "permission denied" in errors.cells[(err_row, 5)][0]
        assert "tail -300" in errors.cells[(err_row, 6)][0]
        # UNKNOWN 清单行：error_code 空、note 注明原因
        unk_rows = [
            row for (row, col) in errors.cells
            if col == 4 and row >= 1 and not errors.cells[(row, col)][0]
        ]
        assert len(unk_rows) == 2
        for row in unk_rows:
            assert errors.cells[(row, 8)][0]  # note 非空


# --------------------------------------------------------------------------
# 8. 实际渲染（xlsxwriter 可用时；openpyxl 只读校验，TD §8）
# --------------------------------------------------------------------------


class TestRenderWorkbook:
    def setup_method(self):
        pytest.importorskip("xlsxwriter")

    def test_three_sheets_and_default_filename(self, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        out = rx.render_xlsx(ok_doc())
        assert out == tmp_path / "insp-20260814120000-node-fx01.xlsx"
        assert out.is_file()
        wb = load_workbook(out)
        assert wb.sheetnames == list(rx.SHEET_NAMES)

    def test_xlsx_out_override(self, tmp_path):
        out = rx.render_xlsx(ok_doc(), tmp_path / "custom" / "my-report.xlsx")
        assert out == tmp_path / "custom" / "my-report.xlsx"
        assert out.is_file()
        # --excel PATH 覆盖语义：不生成默认 <inspection-id>.xlsx
        assert not (tmp_path / "insp-20260814120000-node-fx01.xlsx").exists()

    def test_render_xlsx_file_happy_path(self, tmp_path):
        out = rx.render_xlsx_file(VALID_FIXTURE, tmp_path / "r.xlsx")
        assert out.is_file()
        assert load_workbook(out).sheetnames == list(rx.SHEET_NAMES)

    def test_overview_run_info_and_counts(self, tmp_path):
        docs = rich_docs()
        out = rx.render_xlsx(docs, tmp_path / "r.xlsx")
        ws = load_workbook(out)[rx.SHEET_OVERVIEW]

        # run 信息（RR §3）
        def value_of(label):
            return ws.cell(find_label_row(ws, label), 2).value

        assert value_of("run_id") == _RUN_ID
        assert value_of("inspection_id") == _INSP_ID
        assert value_of("host") == "node-fx01, node-fx02"
        assert value_of("execution_status") == "SUCCESS"
        assert value_of("阈值版本") == "linux-common-p0-v1"
        assert value_of("生成时间")

        # 主机×状态汇总：逐主机计数与 JSON execution_summary 一致
        hdr = find_label_row(ws, "主机×状态汇总") + 1
        for r in (hdr + 1, hdr + 2):
            host = ws.cell(r, 1).value
            counts = {
                s: ws.cell(r, col).value
                for col, s in enumerate(("OK", "WARN", "CRIT", "UNKNOWN"), start=3)
            }
            doc = next(d for d in docs if d["host"]["name"] == host)
            assert counts == rx.count_statuses(doc)
            assert counts["OK"] == doc["execution_summary"]["ok"]

        # 状态计数（合计）：UNKNOWN 独立计数、不混入 OK（RR §3/§5）
        section = find_label_row(ws, "状态计数（合计）")
        ok_v = ws.cell(section + 2, 1).value
        warn_v = ws.cell(section + 2, 2).value
        crit_v = ws.cell(section + 2, 3).value
        unk_v = ws.cell(section + 2, 4).value
        total = ws.cell(section + 2, 5).value
        assert ok_v == 2        # 仅 OK 指标（node-fx01:1 + node-fx02:1）
        assert warn_v == 1
        assert crit_v == 1
        assert unk_v == 3       # UNKNOWN 独立成列，未混入 OK
        assert total == 7
        assert ok_v + warn_v + crit_v + unk_v == total

    def test_local_rows_status_fill_and_crit_red_font(self, tmp_path):
        docs = rich_docs()
        out = rx.render_xlsx(docs, tmp_path / "r.xlsx")
        wb = load_workbook(out)
        ws = wb[rx.SHEET_LOCAL]

        rows = local_rows(ws)
        assert len(rows) == 7  # 每主机每指标一行（5 + 2）
        for row in rows:
            assert row["host"] in ("node-fx01", "node-fx02")
            assert row["metric_id"]
            assert row["status"] in rx.VALID_STATUSES
            assert row["ip"] == "<IP>"
            assert row["threshold_rule"]
            assert row["command"]
            assert "source_anchor" not in row

        # status 单元格：文字 + RR §5 背景色
        for r in range(2, 2 + len(rows)):
            status = ws.cell(r, 8).value
            rgb = ws.cell(r, 8).fill.start_color.rgb
            assert rgb and rgb[-6:] == rx.STATUS_COLORS[status][1:]

        # CRIT 值红色字体（用户需求）：raw_value / normalized_value 列
        crit_rows = [row for row in rows if row["status"] == "CRIT"]
        assert len(crit_rows) == 1
        for r in range(2, 2 + len(rows)):
            if ws.cell(r, 8).value != "CRIT":
                continue
            for col in (5, 6):  # raw_value / normalized_value
                rgb = ws.cell(r, col).font.color.rgb
                assert rgb and rgb[-6:] == "C62828"
                assert ws.cell(r, col).value is not None

    def test_errors_evidence_rows(self, tmp_path):
        docs = rich_docs()
        out = rx.render_xlsx(docs, tmp_path / "r.xlsx")
        ws = load_workbook(out)[rx.SHEET_ERRORS]
        rows = errors_rows(ws)

        # error 非空指标（node-fx01 logs PERMISSION_DENIED）+ UNKNOWN 清单
        # （swap C3、load_1m C5）→ 共 3 行
        assert len(rows) == 3
        err = next(row for row in rows if row["error_code"] == "PERMISSION_DENIED")
        assert err["host"] == "node-fx01"
        assert err["metric_id"] == "local.logs.key_evidence"
        assert err["status"] == "UNKNOWN"
        assert "permission denied" in err["message"]
        assert "tail -300" in err["command"]
        assert err["note"]  # 原因说明非空

        unknowns = [row for row in rows if not row["error_code"]]
        assert len(unknowns) == 2
        assert {row["metric_id"] for row in unknowns} == {
            "local.swap.used_percent", "local.cpu.load_1m",
        }
        for row in unknowns:
            assert row["status"] == "UNKNOWN"
            assert row["note"]  # 文档冲突/缺失原因（RR §3 UNKNOWN 清单）

    def test_local_expands_load_windows_and_filesystem_mounts(self):
        load = make_metric(
            "local.cpu.load_1m", "OK", "0.5", 0.5, unit="数值",
            name="系统负载", rule_id="linux-common-p0-v1.cpu.load.ok",
            value="负载/CPU核数 <= 1.00", ev_command="cat /proc/loadavg; nproc",
            ev_details=[
                {"window": "1 分钟", "load": 0.50, "cpu_cores": 2,
                 "status": "OK", "judgement": "负载/核数=0.25，正常"},
                {"window": "5 分钟", "load": 0.25, "cpu_cores": 2,
                 "status": "OK", "judgement": "负载/核数=0.12，正常"},
                {"window": "15 分钟", "load": 0.10, "cpu_cores": 2,
                 "status": "OK", "judgement": "负载/核数=0.05，正常"},
            ],
        )
        disk = make_metric(
            "local.filesystem.used_percent", "CRIT", "100", 100.0,
            name="磁盘使用率", rule_id="linux-common-p0-v1.filesystem.used_percent.crit",
            value=">85%", ev_command="df -hT",
            ev_details=[
                {"filesystem": "/dev/sda1", "mount": "/", "used_percent": 61, "status": "OK"},
                {"filesystem": "/dev/sdb1", "mount": "/mnt/iso", "used_percent": 100, "status": "CRIT"},
            ],
        )
        inode = make_metric(
            "local.filesystem.inode_used_percent", "OK", "1", 1.0,
            name="inode 使用率", rule_id="linux-common-p0-v1.filesystem.inode_used_percent.ok",
            value="<80%", ev_command="df -i",
            ev_details=[
                {"filesystem": "/dev/sda1", "mount": "/", "used_percent": 3, "status": "OK"},
                {"filesystem": "/dev/sdb1", "mount": "/mnt/iso", "used_percent": 1, "status": "OK"},
            ],
        )
        doc = make_doc("node-detail", [load, disk, inode])

        rows = []
        for metric in doc["metrics"]:
            rows.extend(
                rx._local_row_values(doc, metric, item["detail"])
                for item in rx._metric_rows(metric)
            )

        assert len(rows) == 7
        assert [row["name"] for row in rows[:3]] == [
            "1 分钟系统负载", "5 分钟系统负载", "15 分钟系统负载"
        ]
        assert [row["raw_value"] for row in rows[:3]] == [0.5, 0.25, 0.1]
        assert [row["name"] for row in rows[3:5]] == [
            "磁盘使用率: /", "磁盘使用率: /mnt/iso"
        ]
        assert [row["name"] for row in rows[5:]] == [
            "inode 使用率: /", "inode 使用率: /mnt/iso"
        ]
        assert rows[4]["status"] == "CRIT"
        assert rows[4]["raw_value"] == 100
        assert "挂载点：/mnt/iso" in rows[4]["threshold_rule"]
        assert "判定规则：>85%" in rows[4]["threshold_rule"]
        assert rows[0]["command"] == "cat /proc/loadavg; nproc"
        assert rows[5]["command"] == "df -i"

    def test_rendered_headers_match_contract_constants(self, tmp_path):
        out = rx.render_xlsx(rich_docs(), tmp_path / "r.xlsx")
        wb = load_workbook(out)
        assert [c.value for c in wb[rx.SHEET_LOCAL][1]] == list(rx.LOCAL_HEADERS)
        assert [c.value for c in wb[rx.SHEET_ERRORS][1]] == list(rx.ERRORS_HEADERS)
